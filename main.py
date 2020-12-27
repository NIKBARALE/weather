
import requests
from bs4 import BeautifulSoup
result = requests.get('https://pogoda1.ru/katalog/sverdlovsk-oblast/temperatura-vody/2020')
data = BeautifulSoup(result.content)
temp=data.select('div > script ')
name=data.select(" a ")
name_of_rivers = []
rivers = []
river=[]
months=[]
temp_river=[]
for i in range(1,18):
  foo = str(temp[i])
  el = foo[foo.find('\''):foo.find(';') - 2:].replace('\'','').replace(']','').replace('[','')
  el.split("],[")
  el=el.split(',')
  rivers.append(el)

for i in range(22, 39):
  boo = str(name[i])
  el2 = boo[boo.find('Р'):boo.find('</'):].replace("()",'').strip()
  name_of_rivers.append(el2)

import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
wb.create_sheet(title='2020', index=0 )
wb.remove(wb['Sheet'])
sheet=wb['2020']




#outputfile=open('output.csv','w',newline='')
#outputwriter=csv.writer(outputfile)
months.append('Название реки\Месяц')
river=rivers[0]
for i in range(12):
  months.append(river[i*2])
sheet.append(months)
for i in range(len(rivers)):
  river=rivers[i]
  temp_river.append(name_of_rivers[i])
  for j in range(12):
    temp_river.append(river[2*j+1])
  sheet.append(temp_river)
  temp_river.clear()
wb.save('Temp.xlsx')